from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from core.models import PostResult, ComplianceStatus


def generate_excel_report(posts: list[PostResult]) -> bytes:
    """Genera un reporte Excel con resumen y detalle de errores."""
    wb = Workbook()

    # --- Sheet 1: Resumen ---
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    # Headers
    headers = [
        "URL", "Plataforma", "Estado", "Texto Extraido",
        "Logo Oficial", "Tono", "Puntaje Emotivo",
        "Hashtags Encontrados", "Hashtags Faltantes",
        "Cant. Errores", "Error del Sistema",
    ]
    header_fill = PatternFill(start_color="003DA5", end_color="003DA5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    for col, header in enumerate(headers, 1):
        cell = ws_resumen.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    cumple_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    no_cumple_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    for row, post in enumerate(posts, 2):
        ws_resumen.cell(row=row, column=1, value=post.url)
        ws_resumen.cell(row=row, column=2, value=post.platform.value.capitalize())
        status_cell = ws_resumen.cell(row=row, column=3, value=post.status.value)
        ws_resumen.cell(row=row, column=4, value=post.extracted_text[:300] if post.extracted_text else "")

        if post.analysis:
            a = post.analysis
            ws_resumen.cell(row=row, column=5, value="Si" if a.brand_identity else "No")
            ws_resumen.cell(row=row, column=6, value=a.tone_label)
            ws_resumen.cell(row=row, column=7, value=round(a.emotional_score, 2))
            ws_resumen.cell(row=row, column=8, value=", ".join(a.hashtags_present))
            ws_resumen.cell(row=row, column=9, value=", ".join(a.hashtags_missing))
            ws_resumen.cell(row=row, column=10, value=len(a.design_errors) + len(a.common_errors))
        else:
            ws_resumen.cell(row=row, column=10, value=0)

        ws_resumen.cell(row=row, column=11, value=post.error_message)

        # Colorear fila segun estado
        if post.status == ComplianceStatus.CUMPLE:
            for c in range(1, len(headers) + 1):
                ws_resumen.cell(row=row, column=c).fill = cumple_fill
        elif post.status == ComplianceStatus.NO_CUMPLE:
            for c in range(1, len(headers) + 1):
                ws_resumen.cell(row=row, column=c).fill = no_cumple_fill

    # Ajustar anchos
    ws_resumen.column_dimensions["A"].width = 50
    ws_resumen.column_dimensions["B"].width = 12
    ws_resumen.column_dimensions["C"].width = 12
    ws_resumen.column_dimensions["D"].width = 40
    ws_resumen.column_dimensions["E"].width = 12
    ws_resumen.column_dimensions["H"].width = 25
    ws_resumen.column_dimensions["I"].width = 25

    # --- Sheet 2: Detalle de Errores ---
    ws_errores = wb.create_sheet("Detalle Errores")

    err_headers = ["URL", "Plataforma", "Tipo Error", "Descripcion"]
    for col, header in enumerate(err_headers, 1):
        cell = ws_errores.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    err_row = 2
    for post in posts:
        if post.analysis:
            for err in post.analysis.design_errors:
                ws_errores.cell(row=err_row, column=1, value=post.url)
                ws_errores.cell(row=err_row, column=2, value=post.platform.value)
                ws_errores.cell(row=err_row, column=3, value="Diseno")
                ws_errores.cell(row=err_row, column=4, value=err)
                err_row += 1
            for err in post.analysis.common_errors:
                ws_errores.cell(row=err_row, column=1, value=post.url)
                ws_errores.cell(row=err_row, column=2, value=post.platform.value)
                ws_errores.cell(row=err_row, column=3, value="Comunicacion")
                ws_errores.cell(row=err_row, column=4, value=err)
                err_row += 1

    ws_errores.column_dimensions["A"].width = 50
    ws_errores.column_dimensions["B"].width = 12
    ws_errores.column_dimensions["C"].width = 15
    ws_errores.column_dimensions["D"].width = 60

    # Retornar como bytes
    output = BytesIO()
    wb.save(output)
    return output.getvalue()
